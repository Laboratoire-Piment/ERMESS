# -*- coding: utf-8 -*-
"""
Created on Mon Apr 13 14:32:51 2026

@author: JoPHOBEA
"""

import openpyxl
from openpyxl.chart.layout import Layout,ManualLayout
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series



def _apply_common_style(chart):
    chart.height = 7.5
    chart.width = 20
    chart.legend.overlay = False

    if hasattr(chart, "x_axis"):
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.delete = False

    if hasattr(chart, "y_axis"):
        chart.y_axis.delete = False


# =========================
# HELPERS
# =========================

def _resolve(spec, C):
    return spec(C) if callable(spec) else spec


def _resolve_dynamic(val, C):
    if isinstance(val, str):
        return eval(val, {}, {
            "n_bits": C.time.n_bits,
            "n_store": C.storage.n_store
        })
    return val


def _create_chart(chart_type):
    return {
        "bar": BarChart,
        "line": LineChart,
        "scatter": ScatterChart
    }[chart_type]()


# =========================
# TIMESERIES HANDLER
# =========================

def _build_timeseries_chart(ws, cfg, C):

    # X axis
    if "xvalues" in cfg:
        xmin, ymin, xmax, ymax = cfg["xvalues"]
        ymax = _resolve_dynamic(ymax, C)

        xvalues = Reference(ws, min_col=xmin, min_row=ymin,
                            max_col=xmax, max_row=ymax)
    else:
        xvalues = Reference(ws, min_col=1, min_row=1,
                            max_col=1, max_row=C.time.n_bits + 1)

    # get columns
    cols = []       

    if "custom_series" in cfg:
        cols = [_resolve_dynamic(c, C) for c in cfg["custom_series"]]
     
    else:
        start = cfg["series_col_start"]

        if callable(start):
            cols = start(C)
        else:
            cols = [start + i for i in range(C.storage.n_store)]

    series_len = cfg["series_len"](C) if callable(cfg["series_len"]) else cfg["series_len"]
    # type choice : scatter or line
 #   if len(cols) == 1:
 #       chart = LineChart()

 #       col = cols[0]
 #       values = Reference(ws, min_col=col, min_row=2,
 #                          max_col=col, max_row=series_len+1)

 #       title = ws.cell(row=1, column=col).value or f"Series {col}"
 #       s = Series(values, xvalues, title=title)
        
 #       chart.series.append(s)
        
 #       chart.varyColors = False


 #   else:
    chart = ScatterChart()
        
    for col in cols:
            values = Reference(ws, min_col=col, min_row=2,
                               max_col=col, max_row=series_len+1)
            title = ws.cell(row=1, column=col).value or f"Series {col}"
            s = Series(values, xvalues, title=title)
            if "no smooth" in cfg:
                s.smooth = False

            # Improve readability
            if "thin lines" in cfg:
                s.graphicalProperties.line.width = 12000

            chart.series.append(s)
    if len(cols) == 1:
        empty_values = Reference(ws, min_col=100, min_row=2,
                           max_col=100, max_row=series_len+1)
        empty_title = ""
        empty_s = Series(empty_values, xvalues, title=empty_title)

        chart.series.append(empty_s)
        
            
    chart.scatterStyle = "lineMarker"

    return chart


# =========================
# MAIN FUNCTION
# =========================

def add_excel_charts(Contexte, charts_config):

    file_name = Contexte.postprocess_config.file_name
    wb = openpyxl.load_workbook(file_name)

    for cfg in charts_config:

        ws = wb[cfg["sheet"]]

        # =========================
        # TIMESERIES SPECIAL CASE
        # =========================
        if cfg["type"] == "series of lines":
            chart = _build_timeseries_chart(ws, cfg, Contexte)

        # =========================
        # STANDARD CHARTS
        # =========================
        else:
            chart = _create_chart(cfg["type"])
            
            if cfg.get("stacked"):
                chart.grouping = "stacked"
                chart.overlap = 100

            # ----- DATA
            if "data" in cfg:
                min_col, min_row, max_col, max_row = _resolve(cfg["data"], Contexte)

                if max_row is None:
                    max_row = ws.max_row
                
                

                data = Reference(ws, min_col=min_col, min_row=min_row,max_col=max_col, max_row=max_row)

                chart.add_data(data, titles_from_data=True)

            # ----- CATEGORIES
            if "categories" in cfg:
                cmin_col, cmin_row, cmax_col, cmax_row = _resolve(cfg["categories"], Contexte)

                if cmax_row is None:
                    cmax_row = ws.max_row

                cats = Reference(ws, min_col=cmin_col, min_row=cmin_row,
                                 max_col=cmax_col, max_row=cmax_row)

                chart.set_categories(cats)

        # =========================
        # TITLES
        # =========================
        chart.title = cfg.get("title", "")

        if hasattr(chart, "x_axis"):
            chart.x_axis.title = cfg.get("x_title", "")
            chart.x_axis.delete = False

        if hasattr(chart, "y_axis"):
            chart.y_axis.title = cfg.get("y_title", "")
            chart.y_axis.delete = False

        # =========================
        # AXIS 
        # =========================
        if hasattr(chart, "x_axis"):

            if cfg.get("hide_x_axis"):
                chart.x_axis.delete = True
        
            else:
                chart.x_axis.delete = False
                chart.x_axis.title = cfg.get("x_title", "")
        
                if cfg.get("hide_x_labels"):
                    chart.x_axis.tickLblPos = "none"
            
            if "xmin" in cfg:
                chart.x_axis.scaling.min = cfg["xmin"]
            if "xmax" in cfg:
                chart.x_axis.scaling.max = cfg["xmax"]
        
        if hasattr(chart, "y_axis"):

            if "ymin" in cfg:
                chart.y_axis.scaling.min = cfg["ymin"]

            if "ymax" in cfg:

                if cfg["ymax"] == "auto_energy":
                   
                    chart.y_axis.scaling.max = None
                else:
                    chart.y_axis.scaling.max = cfg["ymax"]

        # =========================
        # STYLE
        # =========================
        chart.height = cfg.get("height", 10)
        chart.width = cfg.get("width", 20)

        if hasattr(chart, "legend"):
            chart.legend.overlay = False
            if "legend_position" in cfg:
                chart.legend.position = cfg["legend_position"]
            if cfg.get("legend") is False:
                chart.legend = None

        # Excel style safe
        if hasattr(chart, "style"):
            chart.style = 2
        
        if "xtick_rotation" in cfg:
            chart.x_axis.tickLblRot = -45
            chart.x_axis.tickLblSkip = 2


        # =========================
        # LAYOUT
        # =========================
        chart.layout = Layout(manualLayout=ManualLayout(x=0.05,y=0.05,w=min(1,max(0.85,chart.width/125)),h=0.75))
        chart.layout.layoutTarget = "inner"
        # =========================
        # ADD TO SHEET
        # =========================        
        ws.add_chart(chart, cfg["position"])

    wb.save(file_name)